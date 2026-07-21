from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from intellibasket.data.ingestion import (
    EXPECTED_HEADERS,
    convertWorkbook,
    normalizeIdentifier,
    normalizeTimestamp,
)


def createWorkbook(workbookPath: Path) -> None:
    workbook = Workbook()
    firstSheet = workbook.active
    firstSheet.title = "Year 2009-2010"
    firstSheet.append(EXPECTED_HEADERS)
    firstSheet.append(
        (
            489434,
            "85048",
            "15CM CHRISTMAS GLASS BALL",
            12,
            datetime(2009, 12, 1, 7, 45),
            6.95,
            13085,
            "United Kingdom",
        )
    )
    firstSheet.append(
        (
            "C489449",
            "22087",
            "PAPER BUNTING",
            -1,
            datetime(2009, 12, 1, 10, 33),
            2.95,
            None,
            "United Kingdom",
        )
    )

    secondSheet = workbook.create_sheet("Year 2010-2011")
    secondSheet.append(EXPECTED_HEADERS)
    secondSheet.append(
        (
            536365,
            "85123A",
            "WHITE HANGING HEART",
            6,
            datetime(2010, 12, 1, 8, 26),
            2.55,
            17850,
            "United Kingdom",
        )
    )
    workbook.save(workbookPath)


def testNormalizeIdentifierRemovesExcelDecimalSuffix() -> None:
    assert normalizeIdentifier(12345.0) == "12345"
    assert normalizeIdentifier("C12345") == "C12345"
    assert normalizeIdentifier(None) == ""


def testNormalizeTimestampUsesHiveFormat() -> None:
    assert normalizeTimestamp(datetime(2026, 7, 21, 8, 30, 15)) == "2026-07-21 08:30:15"


def testConvertWorkbookProducesAuditableCsv(tmp_path: Path) -> None:
    sourcePath = tmp_path / "source.xlsx"
    outputPath = tmp_path / "output.csv"
    createWorkbook(sourcePath)

    ingestionStats = convertWorkbook(sourcePath, outputPath, "TEST_BATCH")

    assert ingestionStats.rowCount == 3
    assert ingestionStats.missingCustomerRows == 1
    assert ingestionStats.cancellationRows == 1
    assert ingestionStats.nonpositiveQuantityRows == 1
    assert len(ingestionStats.sheetStats) == 2

    with outputPath.open(encoding="utf-8", newline="") as outputFile:
        outputRows = list(csv.DictReader(outputFile))

    assert outputRows[0]["invoiceNo"] == "489434"
    assert outputRows[0]["customerId"] == "13085"
    assert outputRows[1]["invoiceNo"] == "C489449"
    assert outputRows[2]["sourceYear"] == "Year 2010-2011"
