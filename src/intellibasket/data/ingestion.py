"""Streaming conversion of Online Retail II into a Hive-friendly CSV file."""

from __future__ import annotations

import csv
import json
from dataclasses import asdict, dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXPECTED_HEADERS = (
    "Invoice",
    "StockCode",
    "Description",
    "Quantity",
    "InvoiceDate",
    "Price",
    "Customer ID",
    "Country",
)

OUTPUT_HEADERS = (
    "sourceYear",
    "invoiceNo",
    "stockCode",
    "description",
    "quantity",
    "invoiceTs",
    "unitPrice",
    "customerId",
    "country",
    "ingestBatchId",
)


@dataclass(slots=True)
class SheetStats:
    """Ingestion counters for one workbook sheet."""

    sourceYear: str
    rowCount: int = 0


@dataclass(slots=True)
class IngestionStats:
    """Auditable output of one workbook conversion run."""

    sourcePath: str
    outputPath: str
    ingestBatchId: str
    startedAt: str
    completedAt: str = ""
    rowCount: int = 0
    missingCustomerRows: int = 0
    cancellationRows: int = 0
    nonpositiveQuantityRows: int = 0
    nonpositivePriceRows: int = 0
    sheetStats: list[SheetStats] = field(default_factory=list)

    def toJson(self) -> str:
        """Serialize stats using API-style camelCase fields."""
        return json.dumps(asdict(self), ensure_ascii=False, indent=2)


def normalizeIdentifier(rawValue: Any) -> str:
    """Normalize Excel identifiers without introducing trailing decimal zeros."""
    if rawValue is None:
        return ""
    if isinstance(rawValue, bool):
        return str(rawValue)
    if isinstance(rawValue, int):
        return str(rawValue)
    if isinstance(rawValue, float) and rawValue.is_integer():
        return str(int(rawValue))
    return str(rawValue).strip()


def normalizeText(rawValue: Any) -> str:
    """Normalize nullable cell text while preserving meaningful punctuation."""
    if rawValue is None:
        return ""
    return " ".join(str(rawValue).strip().split())


def normalizeDecimal(rawValue: Any) -> str:
    """Return a plain decimal string accepted by Hive DECIMAL casts."""
    if rawValue is None or str(rawValue).strip() == "":
        return ""
    try:
        return format(Decimal(str(rawValue)), "f")
    except InvalidOperation as error:
        raise ValueError(f"Invalid decimal value: {rawValue!r}") from error


def normalizeTimestamp(rawValue: Any) -> str:
    """Format Excel timestamps for Hive timestamp conversion."""
    if isinstance(rawValue, datetime):
        return rawValue.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(rawValue, date):
        return datetime.combine(rawValue, datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S")
    if rawValue is None:
        return ""
    return str(rawValue).strip()


def buildBatchId(currentTime: datetime | None = None) -> str:
    """Build a sortable UTC ingestion batch identifier."""
    effectiveTime = currentTime or datetime.now(UTC)
    return effectiveTime.astimezone(UTC).strftime("%Y%m%dT%H%M%SZ")


def validateHeaders(actualHeaders: list[str]) -> None:
    """Fail fast when the workbook schema differs from the documented source."""
    if tuple(actualHeaders) != EXPECTED_HEADERS:
        raise ValueError(
            "Unexpected workbook headers. "
            f"Expected {EXPECTED_HEADERS!r}, received {tuple(actualHeaders)!r}."
        )


def convertWorkbook(
    sourcePath: Path,
    outputPath: Path,
    ingestBatchId: str | None = None,
) -> IngestionStats:
    """Stream both workbook sheets into one standardized UTF-8 CSV file."""
    sourcePath = sourcePath.resolve()
    outputPath = outputPath.resolve()
    if not sourcePath.exists():
        raise FileNotFoundError(f"Source workbook does not exist: {sourcePath}")

    outputPath.parent.mkdir(parents=True, exist_ok=True)
    effectiveBatchId = ingestBatchId or buildBatchId()
    startedAt = datetime.now(UTC).isoformat()
    ingestionStats = IngestionStats(
        sourcePath=str(sourcePath),
        outputPath=str(outputPath),
        ingestBatchId=effectiveBatchId,
        startedAt=startedAt,
    )

    workbook = load_workbook(sourcePath, read_only=True, data_only=True)
    try:
        with outputPath.open("w", encoding="utf-8", newline="") as outputFile:
            csvWriter = csv.writer(outputFile, lineterminator="\n")
            csvWriter.writerow(OUTPUT_HEADERS)

            for worksheet in workbook.worksheets:
                rowIterator = worksheet.iter_rows(values_only=True)
                actualHeaders = [normalizeText(value) for value in next(rowIterator)]
                validateHeaders(actualHeaders)
                columnIndex = {
                    columnName: position for position, columnName in enumerate(actualHeaders)
                }
                currentSheetStats = SheetStats(sourceYear=worksheet.title)

                for sourceRow in rowIterator:
                    invoiceNo = normalizeIdentifier(sourceRow[columnIndex["Invoice"]])
                    quantity = normalizeDecimal(sourceRow[columnIndex["Quantity"]])
                    unitPrice = normalizeDecimal(sourceRow[columnIndex["Price"]])
                    customerId = normalizeIdentifier(sourceRow[columnIndex["Customer ID"]])

                    csvWriter.writerow(
                        (
                            worksheet.title,
                            invoiceNo,
                            normalizeIdentifier(sourceRow[columnIndex["StockCode"]]),
                            normalizeText(sourceRow[columnIndex["Description"]]),
                            quantity,
                            normalizeTimestamp(sourceRow[columnIndex["InvoiceDate"]]),
                            unitPrice,
                            customerId,
                            normalizeText(sourceRow[columnIndex["Country"]]),
                            effectiveBatchId,
                        )
                    )

                    currentSheetStats.rowCount += 1
                    ingestionStats.rowCount += 1
                    if not customerId:
                        ingestionStats.missingCustomerRows += 1
                    if invoiceNo.upper().startswith("C"):
                        ingestionStats.cancellationRows += 1
                    if quantity and Decimal(quantity) <= 0:
                        ingestionStats.nonpositiveQuantityRows += 1
                    if unitPrice and Decimal(unitPrice) <= 0:
                        ingestionStats.nonpositivePriceRows += 1

                ingestionStats.sheetStats.append(currentSheetStats)
    finally:
        workbook.close()

    ingestionStats.completedAt = datetime.now(UTC).isoformat()
    return ingestionStats


def writeIngestionProfile(profilePath: Path, ingestionStats: IngestionStats) -> None:
    """Write a sidecar JSON report for source-to-ODS reconciliation."""
    profilePath = profilePath.resolve()
    profilePath.parent.mkdir(parents=True, exist_ok=True)
    profilePath.write_text(ingestionStats.toJson() + "\n", encoding="utf-8")
