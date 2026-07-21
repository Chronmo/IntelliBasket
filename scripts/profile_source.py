"""Profile the Online Retail II workbook without changing the source file."""

from __future__ import annotations

import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SOURCE_PATH = Path(__file__).resolve().parents[2] / "online+retail+ii" / "online_retail_II.xlsx"


def main() -> None:
    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    profileResult: dict[str, object] = {"source": str(SOURCE_PATH), "sheets": {}}

    totalRows = 0
    totalMissingCustomer = 0
    totalCancellations = 0
    totalNonpositiveQuantity = 0
    totalNonpositivePrice = 0
    invoices: set[str] = set()
    customers: set[str] = set()
    products: set[str] = set()
    validInvoices: set[str] = set()
    validCustomers: set[str] = set()
    validProducts: set[str] = set()
    validRows = 0
    validSalesAmount = 0.0
    countries: Counter[str] = Counter()
    minDate: datetime | None = None
    maxDate: datetime | None = None

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header = [str(value) if value is not None else "" for value in next(rows)]
        columnIndex = {name: position for position, name in enumerate(header)}
        sheetRows = 0
        sheetMissingCustomer = 0
        sheetCancellations = 0

        for row in rows:
            sheetRows += 1
            invoice = row[columnIndex["Invoice"]]
            stockCode = row[columnIndex["StockCode"]]
            quantity = row[columnIndex["Quantity"]]
            invoiceDate = row[columnIndex["InvoiceDate"]]
            price = row[columnIndex["Price"]]
            customerId = row[columnIndex["Customer ID"]]
            country = row[columnIndex["Country"]]

            invoiceText = "" if invoice is None else str(invoice).strip()
            if invoiceText:
                invoices.add(invoiceText)
            if invoiceText.upper().startswith("C"):
                sheetCancellations += 1

            if customerId is None or str(customerId).strip() == "":
                sheetMissingCustomer += 1
            else:
                customers.add(str(customerId).strip())

            if stockCode is not None and str(stockCode).strip():
                products.add(str(stockCode).strip())
            if country is not None and str(country).strip():
                countries[str(country).strip()] += 1
            if quantity is not None and float(quantity) <= 0:
                totalNonpositiveQuantity += 1
            if price is not None and float(price) <= 0:
                totalNonpositivePrice += 1
            if isinstance(invoiceDate, datetime):
                minDate = invoiceDate if minDate is None else min(minDate, invoiceDate)
                maxDate = invoiceDate if maxDate is None else max(maxDate, invoiceDate)

            isValid = (
                bool(invoiceText)
                and not invoiceText.upper().startswith("C")
                and customerId is not None
                and bool(str(customerId).strip())
                and stockCode is not None
                and bool(str(stockCode).strip())
                and isinstance(invoiceDate, datetime)
                and quantity is not None
                and float(quantity) > 0
                and price is not None
                and float(price) > 0
            )
            if isValid:
                validRows += 1
                validInvoices.add(invoiceText)
                validCustomers.add(str(customerId).strip())
                validProducts.add(str(stockCode).strip())
                validSalesAmount += float(quantity) * float(price)

        totalRows += sheetRows
        totalMissingCustomer += sheetMissingCustomer
        totalCancellations += sheetCancellations
        profileResult["sheets"][sheet.title] = {
            "rows": sheetRows,
            "columns": len(header),
            "header": header,
            "missingCustomerRows": sheetMissingCustomer,
            "cancellationRows": sheetCancellations,
        }

    profileResult["overall"] = {
        "rows": totalRows,
        "distinctInvoices": len(invoices),
        "distinctCustomers": len(customers),
        "distinctProducts": len(products),
        "missingCustomerRows": totalMissingCustomer,
        "cancellationRows": totalCancellations,
        "nonpositiveQuantityRows": totalNonpositiveQuantity,
        "nonpositivePriceRows": totalNonpositivePrice,
        "minInvoiceDate": minDate.isoformat(sep=" ") if minDate else None,
        "maxInvoiceDate": maxDate.isoformat(sep=" ") if maxDate else None,
        "topCountriesByRows": countries.most_common(10),
        "validAnalysisRows": validRows,
        "validInvoices": len(validInvoices),
        "validCustomers": len(validCustomers),
        "validProducts": len(validProducts),
        "validSalesAmountGbp": round(validSalesAmount, 2),
        "averageValidBasketAmountGbp": round(validSalesAmount / len(validInvoices), 2)
        if validInvoices
        else 0.0,
    }
    print(json.dumps(profileResult, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
